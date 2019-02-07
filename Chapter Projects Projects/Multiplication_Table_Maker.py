'''
Create a program multiplicationTable.py that takes a number N
from the command line and creates an N×N multiplication table
in an Excel spreadsheet.

For example, when the program is run like this:

py multiplicationTable.py 6

  1 2 3
1 1 2 3
2 2 4 6
3 3 6 9

Row 1 and column A should be used for labels and should be in bold.

'''

import openpyxl, sys, os
import logging
from openpyxl.styles import Font
from openpyxl.utils.cell import get_column_letter,column_index_from_string

logging.basicConfig(level=logging.DEBUG, format= '%(asctime)s - %(levelname)s - %(message)s')

logging.disable(logging.CRITICAL)

if len(sys.argv) < 2:

    #using sys.exit as recommended instead of exit() from research
    sys.exit('\nExiting as no value given!\n\nPlease supply a number when running the script!\n\nExample: multiplicationTable.py 6')

user_num_input = int(sys.argv[1])


#Open Workbook

wb = openpyxl.Workbook()

logging.debug(wb.sheetnames)

current_sheet = wb['Sheet']

# freeze panes so easy to see multiplication sum
current_sheet.freeze_panes= "B2"


title_font = Font(bold=True)


for num in range(1,user_num_input+1):

    cell = current_sheet.cell(row=1, column=num+1)
    cell.font = title_font
    cell.value = num

    cell = current_sheet.cell(row=num+1, column=1)
    cell.font = title_font
    cell.value = num

# slice containing the math area // Is a generator object //
# can create a tuple to visualize


#Start point will always be 'B2'

#End point for Column will be +1 as Col1 is already used // row will be +1 as row1 is used.

table_area = current_sheet['B2':get_column_letter(user_num_input+1) + str(user_num_input+1)]

logging.debug(tuple(table_area)) # tuple with inner tuples // inner tuple for each row in area

for row in table_area:
    for cell in row:
        cell.value = current_sheet.cell(row=cell.row, column=1).value * current_sheet.cell(row=1,column=column_index_from_string(cell.column)).value  #Cell.column returns string letter// needs convert to int

logging.debug(current_sheet.max_row)
logging.debug(current_sheet.max_column)


wb.save('multiplication-table.xlsx')

print('Table is now ready and saved at ' + os.getcwd() + '\nWith the file name "multiplication-table.xlsx"')
