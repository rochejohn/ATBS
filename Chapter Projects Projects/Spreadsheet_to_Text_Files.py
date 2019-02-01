'''
Write a program that performs the tasks of the previous program in reverse order:
The program should open a spreadsheet and write the cells of column A into one text file,
the cells of column B into another text file, and so on.
'''

import openpyxl
from openpyxl.utils import get_column_letter
import logging
logging.basicConfig(level=logging.DEBUG, format = '%(asctime)s - %(levelname)s - %(message)s')
logging.disable(logging.CRITICAL)

wb = openpyxl.load_workbook('files/Text_files_to_spreadsheet.xlsx')
sheet = wb.active

file_num = 0

for num in range(1,sheet.max_column+1):

    file_num +=1
    file_name = 'file' + str(file_num) + '.txt'
    fileobj = open(file_name, 'a')

    column_letter = get_column_letter(num)

    for cellobj in sheet[column_letter]:
        if cellobj.value == None:            # added null as some cells were included with null and crashing script
            continue

        logging.debug(cellobj)
        logging.debug(cellobj.value)

        line = cellobj.value

        fileobj.write(line)

    fileobj.close()

print('\nScript is finished, all column cells have been output to individual files.\n\nFile1, File2, File3 and File4.')