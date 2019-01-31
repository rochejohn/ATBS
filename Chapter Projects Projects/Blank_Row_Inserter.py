'''
Create a program blankRowInserter.py that takes two integers and a filename string
as command line arguments. Let’s call the first integer N and the second integer M.
Starting at row N, the program should insert M blank rows into the spreadsheet.
For example: python blankRowInserter.py 3 2 myProduce.xlsx

You can write this program by reading in the contents of the spreadsheet.
Then, when writing out the new spreadsheet, use a for loop to copy the first N lines.
For the remaining lines, add M to the row number in the output spreadsheet.

'''

import openpyxl,sys,logging,os

logging.basicConfig(level=logging.DEBUG, format= ' %(asctime)s - %(levelname)s - %(message)s')


if len(sys.argv) < 2:
    sys.exit('\nYou need to supply arguments as follows for this script:\n\npy blankRowInserter.py (start row) (number_rows_insert) my_excel_example.xlsx')

start_insert = int(sys.argv[1])
number_rows_insert = int(sys.argv[2])
wb_original = openpyxl.load_workbook(sys.argv[3])


sheet_original = wb_original.active

wb_new = openpyxl.Workbook()
sheet_new = wb_new.active


# write all lines to new workbook upto the blank insert


for row in range(1,start_insert):

    cells = sheet_original[row]

    for cell in cells:

        sheet_new[cell.coordinate].value = sheet_original[cell.coordinate].value



for row in range(sheet_original.max_row,start_insert-1,-1):

    cells = sheet_original[row]

    for col, cellobj in enumerate(cells,1):

        sheet_new.cell(row=row+start_insert, column=col).value = sheet_original[cellobj.coordinate].value



wb_new.save('modified_with_blank_inserts.xlsx')

print('\nFinished!\n\nYou can find your edited file in ' + os.getcwd() +' under the filename "modified_with_blank_inserts.xlsx!"')



