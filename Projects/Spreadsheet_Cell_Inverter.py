#! /usr/bin/env python3


'''
Write a program to invert the row and column of the cells in the spreadsheet.
For example, the value at row 5, column 3 will be at row 3, column 5 (and vice versa).
This should be done for all cells in the spreadsheet.

You can write this program by using nested for loops to read in the spreadsheet’s data into a list of lists data structure.
This data structure could have sheetData[x][y] for the cell at column x and row y.
Then, when writing out the new spreadsheet, use sheetData[y][x] for the cell at column x and row y.
'''


import openpyxl,logging,os
logging.basicConfig(level=logging.DEBUG, format=' %(asctime)s - %(levelname)s - %(message)s')
from openpyxl.utils import get_column_letter, column_index_from_string

logging.disable(logging.CRITICAL)


user_input = input('Please supply a spreadsheet to be inverted!\n\nSpreadsheet name: ')

#user_input = 'Before_inverted.xlsx'

wb_original = openpyxl.load_workbook(user_input)

sheet_original = wb_original.active

wb_new = openpyxl.Workbook()
sheet_new = wb_new.active

for row in range(1,sheet_original.max_row+1):

    for cellobj in sheet_original[row]:

        col_num = column_index_from_string(cellobj.column)
        row_num = cellobj.row


        sheet_new.cell(row=col_num,column=row_num).value = cellobj.value

wb_new.save('Inverted_spreadsheet.xlsx')

print('\nFinished processing!\n\nInverted file saved as "Inverted_spreadsheet.xlsx at ' +os.getcwd() +'!' )


